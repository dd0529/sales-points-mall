import hashlib
import openpyxl

def safe_str(val):
    """Convert cell value to string, treating None as empty."""
    if val is None:
        return ''
    return str(val)

def parse_excel(filepath):
    """Parse Excel file, return list of dicts and any error message."""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
    except Exception as e:
        return None, f'文件无法打开: {str(e)}'

    # Find header row (Row 1 per the sample file)
    header_row = [safe_str(c.value) for c in ws[1]]

    # Identify key columns by name substring matching
    col_map = {}
    for key, labels in [
        ('date', ['日期', '会计日']),
        ('opening_date', ['门店开业时间']),
        ('biz_org_name', ['业务机构名称']),
        ('product_name', ['商品名称']),
        ('salesperson', ['营业员名字']),
        ('quantity', ['销售数量']),
        ('amount', ['实际金额']),
    ]:
        for i, h in enumerate(header_row):
            for label in labels:
                if label in h:
                    if key == 'amount' and '不含税' in h:
                        continue
                    col_map[key] = i
                    break
            if key in col_map:
                break

    required = ['biz_org_name', 'product_name', 'salesperson', 'quantity', 'amount']
    missing = [k for k in required if k not in col_map]
    if missing:
        return None, f'缺少必要列: {", ".join(missing)}'

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        try:
            amount = float(row[col_map['amount']]) if row[col_map['amount']] else 0
        except (ValueError, TypeError):
            amount = 0
        try:
            qty = float(row[col_map['quantity']]) if row[col_map['quantity']] else 0
        except (ValueError, TypeError):
            qty = 0

        if amount <= 0:
            continue

        prod = safe_str(row[col_map['product_name']])
        person = safe_str(row[col_map['salesperson']])
        biz = safe_str(row[col_map['biz_org_name']])

        # Skip total/summary rows (empty product or empty biz with empty person)
        if not prod or (not biz and not person):
            continue

        # Skip 桃花姬牌阿胶核桃芝麻糕
        if prod == '桃花姬牌阿胶核桃芝麻糕':
            continue

        date_val = safe_str(row[col_map['date']]) if 'date' in col_map else ''
        opening = safe_str(row[col_map['opening_date']]) if 'opening_date' in col_map else ''

        hash_str = f'{date_val}|{biz}|{prod}|{person}|{qty}|{amount}'
        row_hash = hashlib.md5(hash_str.encode()).hexdigest()

        records.append({
            'date': date_val,
            'opening_date': opening,
            'biz_org_name': biz,
            'product_name': prod,
            'salesperson': person,
            'quantity': qty,
            'amount': amount,
            'points': round(amount / 100),
            'row_hash': row_hash,
        })

    return records, None
